#!/usr/bin/env python3
"""
Import Kestrel Proxy Observational Data v2 from Excel into ArangoDB observations collection.

Reads all per-satellite sheets from the XLSX attachment, transforms the flat
thermal_anomaly_flag to the nested thermal.anomaly_flag format used by the existing
observation schema, and bulk-inserts records directly into the observations collection.

New fields introduced by this dataset (stored in their schema sub-objects):
  pass_id, frame_index, observation_mode, sensors_active, illumination (top-level)
  attitude: roll_deg, pitch_deg, yaw_deg, stability_flag
  thermal: surface_temp_K, anomaly_flag
  material_signature: reflectivity_index, inferred_material, material_confidence
  proximity_state: range_km, relative_velocity_ms
  maneuver_indicator: delta_v_residual_ms, maneuver_confidence, maneuver_flag
  orbital_decay_indicator: perigee_drift_km_per_day, estimated_perigee_km

Source identifier: kestrel_proxy_v2
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import openpyxl
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

import database.connection as db_conn
from database.connection import COLLECTION_NAME, COLLECTION_OBSERVATIONS
from database.observation_graph_ops import create_edges_for_observation


SUMMARY_SHEET = "Summary"
SOURCE_VALUE = "kestrel_proxy_v2"

FLAT_FIELD_MAP = {
    "norad_id": "norad_id",
    "object_name": "object_name",
    "object_type": "object_type",
    "origin_country": "origin_country",
    "observation_epoch": "observation_epoch",
    "source": "source",
    "pass_id": "pass_id",
    "frame_index": "frame_index",
    "observation_mode": "observation_mode",
    "sensors_active": "sensors_active",
    "illumination": "illumination",
    "estimated_mass_kg": "estimated_mass_kg",
    "spin_rate_rpm": "spin_rate_rpm",
    "derived_health_score": "derived_health_score",
}

ATTITUDE_FIELDS = ["roll_deg", "pitch_deg", "yaw_deg", "stability_flag"]
MATERIAL_FIELDS = ["reflectivity_index", "inferred_material", "material_confidence"]
PROXIMITY_FIELDS = ["range_km", "relative_velocity_ms"]
MANEUVER_FIELDS = ["delta_v_residual_ms", "maneuver_confidence"]
ORBITAL_DECAY_FIELDS = ["perigee_drift_km_per_day", "estimated_perigee_km"]

MANEUVER_FLAG_THRESHOLD_MS = 0.5


def _cast_value(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip().lower() in ("", "nan", "none", "n/a"):
        return None
    return value


def _parse_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes"):
        return True
    if s in ("false", "0", "no"):
        return False
    return None


_MANEUVER_FLAG_TRUE = {"suspected_maneuver", "maneuver_detected", "detected"}
_MANEUVER_FLAG_FALSE = {"no_maneuver", "nominal"}


def _parse_maneuver_flag(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in _MANEUVER_FLAG_TRUE:
        return True
    if s in _MANEUVER_FLAG_FALSE:
        return False
    return _parse_bool(value)


_STABILITY_FLAG_TRUE = {"anomalous", "unstable", "degraded"}
_STABILITY_FLAG_FALSE = {"nominal", "stable"}


def _parse_stability_flag(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in _STABILITY_FLAG_TRUE:
        return True
    if s in _STABILITY_FLAG_FALSE:
        return False
    return _parse_bool(value)


def read_sheets(xlsx_path: str) -> list[dict]:
    """
    Read all non-Summary sheets from the Excel file and return a flat list of
    observation dicts ready for insertion.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        if sheet_name == SUMMARY_SHEET:
            continue

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [h for h in next(rows, [])]

        if not headers:
            continue

        sheet_count = 0
        for row in rows:
            raw = dict(zip(headers, row))

            doc = {}
            for col, field in FLAT_FIELD_MAP.items():
                val = _cast_value(raw.get(col))
                if val is not None:
                    doc[field] = val

            thermal_flag = _parse_bool(raw.get("thermal_anomaly_flag"))
            if thermal_flag is not None:
                doc["thermal"] = {"anomaly_flag": thermal_flag}

            attitude = {f: _cast_value(raw.get(f)) for f in ATTITUDE_FIELDS if f != "stability_flag" and _cast_value(raw.get(f)) is not None}
            stability_flag = _parse_stability_flag(raw.get("stability_flag"))
            if stability_flag is not None:
                attitude["stability_flag"] = stability_flag
            if attitude:
                doc["attitude"] = attitude

            material_signature = {f: _cast_value(raw.get(f)) for f in MATERIAL_FIELDS if _cast_value(raw.get(f)) is not None}
            if material_signature:
                doc["material_signature"] = material_signature

            proximity_state = {f: _cast_value(raw.get(f)) for f in PROXIMITY_FIELDS if _cast_value(raw.get(f)) is not None}
            if proximity_state:
                doc["proximity_state"] = proximity_state

            maneuver_indicator = {f: _cast_value(raw.get(f)) for f in MANEUVER_FIELDS if _cast_value(raw.get(f)) is not None}
            maneuver_flag = _parse_maneuver_flag(raw.get("maneuver_flag"))
            if maneuver_flag is None:
                dv = maneuver_indicator.get("delta_v_residual_ms")
                if dv is not None:
                    try:
                        maneuver_flag = float(dv) >= MANEUVER_FLAG_THRESHOLD_MS
                    except (TypeError, ValueError):
                        pass
            if maneuver_flag is not None:
                maneuver_indicator["maneuver_flag"] = maneuver_flag
            if maneuver_indicator:
                doc["maneuver_indicator"] = maneuver_indicator

            orbital_decay_indicator = {f: _cast_value(raw.get(f)) for f in ORBITAL_DECAY_FIELDS if _cast_value(raw.get(f)) is not None}
            if orbital_decay_indicator:
                doc["orbital_decay_indicator"] = orbital_decay_indicator

            surface_temp = _cast_value(raw.get("surface_temp_K"))
            if surface_temp is not None:
                doc.setdefault("thermal", {})["surface_temp_K"] = surface_temp

            if "norad_id" not in doc or "observation_epoch" not in doc:
                continue

            doc.setdefault("source", SOURCE_VALUE)

            all_records.append(doc)
            sheet_count += 1

        print(f"  Sheet '{sheet_name}': {sheet_count} records", flush=True)

    wb.close()
    return all_records


def enable_observations_for_norad_ids(norad_ids: set[int], dry_run: bool) -> dict:
    """Set observations_enabled=true on satellite documents for each NORAD ID."""
    enabled = 0
    not_found = 0

    for norad_id in sorted(norad_ids):
        cursor = db_conn.db.aql.execute(
            """
            FOR sat IN @@satellites
                FILTER sat.canonical.norad_cat_id == @norad_id
                LIMIT 1
                RETURN sat._key
            """,
            bind_vars={"@satellites": COLLECTION_NAME, "norad_id": norad_id},
        )
        keys = list(cursor)
        if not keys:
            not_found += 1
            if not dry_run:
                print(f"  [WARN] Satellite NORAD {norad_id} not found in satellites collection", flush=True)
            continue

        if not dry_run:
            db_conn.db.aql.execute(
                """
                FOR sat IN @@satellites
                    FILTER sat.canonical.norad_cat_id == @norad_id
                    UPDATE sat WITH {canonical: MERGE(sat.canonical, {observations_enabled: true})} IN @@satellites
                """,
                bind_vars={"@satellites": COLLECTION_NAME, "norad_id": norad_id},
            )
        enabled += 1

    return {"enabled": enabled, "not_found": not_found}


def build_existing_set(norad_ids: list[int]) -> set[tuple]:
    """Fetch (norad_id, observation_epoch, source) tuples already in DB for dedup."""
    cursor = db_conn.db.aql.execute(
        """
        FOR obs IN @@observations
            FILTER obs.norad_id IN @norad_ids AND obs.source == @source
            RETURN {norad_id: obs.norad_id, observation_epoch: obs.observation_epoch, source: obs.source}
        """,
        bind_vars={
            "@observations": COLLECTION_OBSERVATIONS,
            "norad_ids": norad_ids,
            "source": SOURCE_VALUE,
        },
    )
    return {(r["norad_id"], r["observation_epoch"], r["source"]) for r in cursor}


def import_records(records: list[dict], batch_size: int, dry_run: bool, create_edges: bool) -> dict:
    """Insert observation records into ArangoDB in batches."""
    if dry_run:
        print(f"  [DRY RUN] Would insert up to {len(records)} observations (skipping DB check)", flush=True)
        return {"inserted": len(records), "skipped_duplicates": 0, "errors": 0}

    obs_col = db_conn.db.collection(COLLECTION_OBSERVATIONS)

    norad_ids = list({r["norad_id"] for r in records})
    print(f"Checking existing observations for {len(norad_ids)} NORAD IDs...", flush=True)
    existing_set = build_existing_set(norad_ids)
    print(f"Found {len(existing_set)} existing kestrel_proxy_v2 observations (will skip duplicates)", flush=True)

    inserted = 0
    skipped = 0
    errors = 0

    batch = []
    for rec in records:
        key = (rec["norad_id"], rec["observation_epoch"], rec.get("source", SOURCE_VALUE))
        if key in existing_set:
            skipped += 1
            continue

        batch.append(rec)
        existing_set.add(key)

        if len(batch) >= batch_size:
            if not dry_run:
                _flush_batch(obs_col, batch, create_edges)
            inserted += len(batch)
            batch = []

            if inserted % 1000 == 0:
                print(f"  Progress: {inserted} inserted, {skipped} skipped duplicates", flush=True)

    if batch:
        if not dry_run:
            _flush_batch(obs_col, batch, create_edges)
        inserted += len(batch)

    return {"inserted": inserted, "skipped_duplicates": skipped, "errors": errors}


def _flush_batch(obs_col, batch: list[dict], create_edges: bool):
    obs_col.import_bulk(batch, on_duplicate="ignore")
    if create_edges:
        cursor = db_conn.db.aql.execute(
            """
            FOR obs IN @@observations
                FILTER obs.norad_id IN @norad_ids
                    AND obs.observation_epoch IN @epochs
                    AND obs.source == @source
                RETURN obs
            """,
            bind_vars={
                "@observations": COLLECTION_OBSERVATIONS,
                "norad_ids": [r["norad_id"] for r in batch],
                "epochs": [r["observation_epoch"] for r in batch],
                "source": SOURCE_VALUE,
            },
        )
        for obs_doc in cursor:
            try:
                create_edges_for_observation(obs_doc)
            except Exception:
                pass


def run(xlsx_path: str, batch_size: int = 500, dry_run: bool = False, skip_edges: bool = False):
    print(f"Kestrel Proxy v2 Import — {'DRY RUN' if dry_run else 'LIVE'}", flush=True)
    print(f"Source file: {xlsx_path}", flush=True)
    print(flush=True)

    if not dry_run:
        if not db_conn.connect_arangodb():
            print("Failed to connect to ArangoDB")
            sys.exit(1)

    print("Reading Excel sheets...", flush=True)
    records = read_sheets(xlsx_path)
    print(f"Total records read: {len(records)}", flush=True)
    print(flush=True)

    norad_ids = {r["norad_id"] for r in records}
    print(f"Enabling observations for {len(norad_ids)} NORAD IDs...", flush=True)
    if not dry_run:
        enable_result = enable_observations_for_norad_ids(norad_ids, dry_run=False)
        print(f"  observations_enabled set: {enable_result['enabled']}, satellites not found: {enable_result['not_found']}", flush=True)
    else:
        print(f"  [DRY RUN] Would enable observations for NORAD IDs: {sorted(norad_ids)}", flush=True)
    print(flush=True)

    print(f"Importing {len(records)} observations (batch_size={batch_size})...", flush=True)
    stats = import_records(records, batch_size=batch_size, dry_run=dry_run, create_edges=not skip_edges)
    print(flush=True)

    print("Import complete!", flush=True)
    print(f"  Inserted:           {stats['inserted']}", flush=True)
    print(f"  Skipped (dup):      {stats['skipped_duplicates']}", flush=True)
    print(f"  Errors:             {stats['errors']}", flush=True)
    print(f"  Total submitted:    {len(records)}", flush=True)


def main():
    default_xlsx = str(
        Path(__file__).parent.parent.parent
        / ".zenflow-attachments"
        / "cbcb70b2-0747-49df-b0f8-51db7a6ead87.xlsx"
    )

    parser = argparse.ArgumentParser(description="Import Kestrel Proxy v2 observational data")
    parser.add_argument(
        "--file",
        default=default_xlsx,
        help="Path to the Kestrel Proxy v2 Excel file (default: attachment in repo)",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=500,
        help="Records per batch insert (default: 500)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and count records without writing to the database",
    )
    parser.add_argument(
        "--skip-edges",
        action="store_true",
        help="Skip graph edge creation after insert (faster, run populate_all_observation_edges separately)",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.file):
        print(f"Excel file not found: {args.file}")
        sys.exit(1)

    run(
        xlsx_path=args.file,
        batch_size=args.batch_size,
        dry_run=args.dry_run,
        skip_edges=args.skip_edges,
    )


if __name__ == "__main__":
    main()
